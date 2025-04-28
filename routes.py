import os
import logging
import json
import base64
from datetime import datetime
from flask import render_template, url_for, flash, redirect, request, jsonify, session
from flask_login import login_user, logout_user, current_user, login_required
from werkzeug.utils import secure_filename
from flask_wtf.csrf import generate_csrf
from app import app, db
from models import User, ExtractedForm
from forms import LoginForm, RegistrationForm, TemplateSelectionForm, FormUploadForm
from gemini_form_extractor import GeminiFormExtractor
from form_templates import FORM_TEMPLATES
import tempfile

# Set up logging
logger = logging.getLogger(__name__)

# CSRF token context processor
@app.context_processor
def inject_csrf_token():
    return dict(csrf_token=generate_csrf())

@app.route('/')
@app.route('/index')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user is None or not user.check_password(form.password.data):
            flash('Invalid username or password', 'danger')
            return redirect(url_for('login'))
        
        login_user(user, remember=form.remember_me.data)
        flash('Login successful!', 'success')
        
        next_page = request.args.get('next')
        if not next_page or not next_page.startswith('/'):
            next_page = url_for('dashboard')
        return redirect(next_page)
    
    return render_template('login.html', title='Sign In', form=form)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = RegistrationForm()
    if form.validate_on_submit():
        user = User(username=form.username.data, email=form.email.data)
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        
        flash('Registration successful! You can now log in.', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html', title='Register', form=form)

@app.route('/logout')
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    # Get all the user's extracted forms, ordered by newest first
    extracted_forms = ExtractedForm.query.filter_by(user_id=current_user.id).order_by(ExtractedForm.created_at.desc()).all()
    return render_template('dashboard.html', title='Dashboard', extracted_forms=extracted_forms)

@app.route('/template-selection', methods=['GET', 'POST'])
@login_required
def template_selection():
    form = TemplateSelectionForm()
    if form.validate_on_submit():
        # Store the selected template in the session
        session['selected_template'] = form.template.data
        return redirect(url_for('form_upload'))
    
    return render_template('template_selection.html', title='Select Template', form=form)

@app.route('/form-upload', methods=['GET', 'POST'])
@login_required
def form_upload():
    # Check if template was selected
    if 'selected_template' not in session:
        flash('Please select a template first', 'warning')
        return redirect(url_for('template_selection'))
    
    selected_template = session['selected_template']
    form = FormUploadForm()
    
    if form.validate_on_submit() or ('camera_image' in request.form and request.form['camera_image']):
        temp_dir = tempfile.mkdtemp()
        temp_path = ""
        filename = ""
        
        try:
            # Check if we received camera data or file upload
            if 'camera_image' in request.form and request.form['camera_image']:
                try:
                    # Handle camera image (base64 encoded)
                    camera_data = request.form['camera_image']
                    
                    # Log data length for debugging
                    logger.debug(f"Received camera data length: {len(camera_data)}")
                    
                    # Remove the data URL prefix if present (e.g., "data:image/jpeg;base64,")
                    if ',' in camera_data:
                        logger.debug("Data URL prefix detected, splitting at comma")
                        camera_data = camera_data.split(',', 1)[1]
                    else:
                        logger.debug("No data URL prefix detected")
                    
                    # Log data after prefix is removed
                    logger.debug(f"Camera data length after prefix removal: {len(camera_data)}")
                    
                    # Decode base64 data
                    try:
                        image_data = base64.b64decode(camera_data)
                        logger.debug(f"Successfully decoded base64 data, length: {len(image_data)} bytes")
                    except Exception as decode_error:
                        logger.error(f"Error decoding base64 data: {str(decode_error)}")
                        raise
                    
                    # Save to temp file
                    filename = f"camera_capture_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                    temp_path = os.path.join(temp_dir, filename)
                    
                    with open(temp_path, 'wb') as f:
                        f.write(image_data)
                    
                    logger.debug(f"Saved camera capture to {temp_path}")
                except Exception as camera_error:
                    logger.error(f"Error processing camera image: {str(camera_error)}")
                    flash(f"Error processing camera image: {str(camera_error)}", 'danger')
                    return render_template('form_upload.html', title='Upload Form', form=form, template=selected_template)
                
            elif form.form_file.data:
                # Handle regular file upload
                uploaded_file = form.form_file.data
                filename = secure_filename(uploaded_file.filename)
                temp_path = os.path.join(temp_dir, filename)
                uploaded_file.save(temp_path)
                logger.debug(f"Saved uploaded file to {temp_path}")
            else:
                flash('No file or camera image provided', 'danger')
                return render_template('form_upload.html', title='Upload Form', form=form, template=selected_template)
            
            # Extract data using Google Gemini API
            api_key = app.config['GOOGLE_API_KEY']
            
            logger.debug(f"Using Google Gemini API for extraction")
            logger.debug(f"API key length: {len(api_key) if api_key else 0}")
            
            extractor = GeminiFormExtractor(api_key=api_key)
            
            extracted_data = extractor.extract_form_data(temp_path, selected_template)
            
            # Store extracted data in session to be used in review page
            session['extracted_data'] = extracted_data
            session['uploaded_filename'] = filename
            
            flash('Form data extracted successfully!', 'success')
            return redirect(url_for('review_data'))
            
        except Exception as e:
            logger.error(f"Error extracting form data: {str(e)}")
            flash(f'Error extracting form data: {str(e)}', 'danger')
        finally:
            # Clean up the temporary file
            if os.path.exists(temp_path):
                os.remove(temp_path)
            if os.path.exists(temp_dir):
                os.rmdir(temp_dir)
    
    return render_template('form_upload.html', title='Upload Form', form=form, template=selected_template)

@app.route('/review-data', methods=['GET', 'POST'])
@login_required
def review_data():
    # Check if extracted data is available in session
    if 'extracted_data' not in session or 'selected_template' not in session:
        flash('No extracted data to review. Please upload a form first.', 'warning')
        return redirect(url_for('template_selection'))
    
    # Get the template structure for the selected template
    template_type = session['selected_template']
    template = FORM_TEMPLATES.get(template_type, {})
    
    if request.method == 'POST':
        # Debug: Log all form data received
        logger.debug("FORM DATA RECEIVED:")
        for key, value in request.form.items():
            logger.debug(f"  {key}: {value}")
            
        # Get the updated form data from the submitted form
        updated_data = {}
        
        # Process form data based on the template structure
        for section_name, fields_info in template.items():
            section_data = {}
            
            # Go through each field in this section template
            for field_info in fields_info:
                field_name = field_info["field"]
                
                # Match the field ID format used in the template
                # In the template: {{ section }}_{{ field|replace(' ', '_')|replace("'", "") }}
                field_id = f"{section_name}_{field_name}".replace(" ", "_").replace("'", "")
                
                # Try both formats - with and without space replacement
                field_value = request.form.get(field_id, "")
                
                # If empty, try with spaces (as seen in the form submission)
                if not field_value:
                    alternate_field_id = f"{section_name} {field_name}"
                    field_value = request.form.get(alternate_field_id, "")
                    if field_value:
                        logger.debug(f"Found value using alternate ID: {alternate_field_id}")
                
                # Debug this specific field
                logger.debug(f"Field: {field_id}, Value from form: '{field_value}'")
                
                # Store the form value 
                section_data[field_name] = field_value
                
            updated_data[section_name] = section_data
            
        # Log data before saving
        logger.debug(f"Saving updated form data: {updated_data}")
        
        # Save the extracted form data to the database
        extracted_form = ExtractedForm(
            user_id=current_user.id,
            template_type=session['selected_template'],
            file_name=session['uploaded_filename']
        )
        extracted_form.set_data(updated_data)
        
        db.session.add(extracted_form)
        db.session.commit()
        
        # Clear the session data
        session.pop('extracted_data', None)
        session.pop('selected_template', None)
        session.pop('uploaded_filename', None)
        
        flash('Form data saved successfully!', 'success')
        return redirect(url_for('dashboard'))
    
    # For GET requests, pass the template data to the template
    extracted_data = session['extracted_data']
    logger.debug(f"Displaying data for review: {extracted_data}")
    
    return render_template(
        'review_data.html', 
        title='Review Data', 
        template_type=session['selected_template'],
        extracted_data=extracted_data
    )

@app.route('/view-form/<int:form_id>')
@login_required
def view_form(form_id):
    # Get the form data from the database
    extracted_form = ExtractedForm.query.get_or_404(form_id)
    
    # Check if the form belongs to the current user
    if extracted_form.user_id != current_user.id:
        flash('You do not have permission to view this form.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get form data and ensure it's in the proper format for the template
    form_data = extracted_form.get_data()
    
    # Log form data for debugging
    logger.debug(f"Original form data retrieved: {form_data}")
    
    # Ensure all sections are dictionaries for consistency
    # Add missing fields from template if not present
    template_type = extracted_form.template_type
    template = FORM_TEMPLATES.get(template_type, {})
    
    # Create a complete form structure with empty values if fields are missing
    complete_form_data = {}
    
    for section_name, fields_info in template.items():
        section_data = {}
        # Initialize with empty data for all fields from template
        for field_info in fields_info:
            field_name = field_info["field"]
            section_data[field_name] = ""
            
        # If section exists in saved data, use those values
        if section_name in form_data and isinstance(form_data[section_name], dict):
            for field_name, value in form_data[section_name].items():
                section_data[field_name] = value
                
        complete_form_data[section_name] = section_data
    
    logger.debug(f"Complete form data with template fields: {complete_form_data}")
    
    return render_template(
        'review_data.html', 
        title='View Form', 
        template_type=extracted_form.template_type,
        extracted_data=complete_form_data,
        view_only=True,
        form_id=form_id
    )

@app.route('/export-form/<int:form_id>/<format>')
@login_required
def export_form(form_id, format):
    # Get the form data from the database
    extracted_form = ExtractedForm.query.get_or_404(form_id)
    
    # Check if the form belongs to the current user
    if extracted_form.user_id != current_user.id:
        flash('You do not have permission to export this form.', 'danger')
        return redirect(url_for('dashboard'))
    
    # For debugging: Print raw data from database
    raw_data = extracted_form.extracted_data
    logger.debug(f"Raw data from DB: {raw_data}")
    
    # Get the form data
    form_data = extracted_form.get_data()
    logger.debug(f"Parsed JSON data: {form_data}")
    
    template_type = extracted_form.template_type
    logger.debug(f"Template type: {template_type}")
    
    # Ensure all sections are dictionaries for consistency
    # Add missing fields from template if not present
    template = FORM_TEMPLATES.get(template_type, {})
    
    # Create a complete form structure with empty values if fields are missing
    complete_form_data = {}
    
    # Debug: Log template structure
    logger.debug(f"Template structure: {template.keys()}")
    
    for section_name, fields_info in template.items():
        section_data = {}
        logger.debug(f"Processing section: {section_name} with {len(fields_info)} fields")
        
        # Initialize with empty data for all fields from template
        for field_info in fields_info:
            field_name = field_info["field"]
            section_data[field_name] = ""
            logger.debug(f"Initialized field: {field_name}")
            
        # If section exists in saved data, use those values
        if section_name in form_data and isinstance(form_data[section_name], dict):
            logger.debug(f"Found section {section_name} in form data with {len(form_data[section_name])} fields")
            for field_name, value in form_data[section_name].items():
                section_data[field_name] = value
                logger.debug(f"Set field {field_name} = '{value}'")
        else:
            logger.debug(f"Section {section_name} not found in form data or not a dictionary")
                
        complete_form_data[section_name] = section_data
    
    # Final debug: Log complete data being sent
    logger.debug(f"Complete data being sent: {complete_form_data}")
    
    # Return the data to be handled by client-side export functions
    if format == 'json':
        # For Excel export, we'll use json as an intermediate
        data = {
            'formId': form_id,
            'templateType': template_type,
            'fileName': extracted_form.file_name,
            'extractedData': complete_form_data
        }
        logger.debug(f"JSON response keys: {data.keys()}")
        logger.debug(f"JSON extractedData sections: {data['extractedData'].keys()}")
        logger.debug(f"Sample data from first section: {list(data['extractedData'].values())[0] if data['extractedData'] else 'No data'}")
        
        return jsonify(data)
    elif format == 'excel':
        # Server-side Excel generation
        import io
        from flask import send_file
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Debugging information
        logger.debug(f"Generating Excel for form {form_id} - Template: {template_type}")
        logger.debug(f"Number of sections: {len(complete_form_data)}")
        for section_name, fields in complete_form_data.items():
            logger.debug(f"Section {section_name} has {len(fields)} fields")
            for field, value in fields.items():
                logger.debug(f"  {field}: {value}")
        
        # Create an in-memory output file
        output = io.BytesIO()
        
        # Create Excel workbook
        workbook = openpyxl.Workbook()
        
        # Style elements
        title_font = Font(size=16, bold=True)
        header_font = Font(size=12, bold=True)
        section_font = Font(size=14, bold=True)
        normal_font = Font(size=11)
        
        # Header fill
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alignment
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # Create summary sheet
        summary_sheet = workbook.active
        summary_sheet.title = 'Summary'
        
        # Add title with formatting
        summary_sheet['A1'] = f"{template_type} Form"
        summary_sheet['A1'].font = title_font
        
        # Add form details
        summary_sheet['A2'] = "Original File:"
        summary_sheet['B2'] = extracted_form.file_name
        
        summary_sheet['A3'] = "Export Date:"
        summary_sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Add form sections header
        summary_sheet['A5'] = "Form Sections:"
        summary_sheet['A5'].font = header_font
        
        # List all sections
        row = 6
        for idx, section_name in enumerate(complete_form_data.keys()):
            summary_sheet[f'A{row}'] = f"{idx + 1}. {section_name}"
            row += 1
        
        # Set column widths
        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 40
        
        # For each section, create a detailed sheet with all fields and values
        for section_name, fields in complete_form_data.items():
            # Skip empty sections
            if not fields:
                continue
                
            # Create sheet name (ensure valid Excel sheet name)
            safe_section_name = section_name.replace('/', '-').replace('\\', '-')[:31]
            
            try:
                # Create sheet for this section
                section_sheet = workbook.create_sheet(title=safe_section_name)
                
                # Add section title
                section_sheet['A1'] = section_name
                section_sheet['A1'].font = section_font
                section_sheet.merge_cells('A1:B1')
                
                # Add field headers
                section_sheet['A3'] = "Field"
                section_sheet['B3'] = "Value"
                
                # Style headers
                for cell in section_sheet['3:3']:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = center_align
                
                # Add data rows
                row = 4
                for field, value in fields.items():
                    # Add field name
                    section_sheet[f'A{row}'] = field
                    section_sheet[f'A{row}'].border = thin_border
                    section_sheet[f'A{row}'].alignment = left_align
                    
                    # Add field value (handle None values)
                    section_sheet[f'B{row}'] = value if value is not None else ''
                    section_sheet[f'B{row}'].border = thin_border
                    section_sheet[f'B{row}'].alignment = left_align
                    
                    row += 1
                
                # Set column widths
                section_sheet.column_dimensions['A'].width = 30
                section_sheet.column_dimensions['B'].width = 50
                
                logger.debug(f"Added sheet {safe_section_name} with {len(fields)} fields")
            except Exception as e:
                logger.error(f"Error creating sheet {safe_section_name}: {str(e)}")
                # Continue with next section on error
        
        # Create all data sheet (single table with all sections/fields)
        all_data_sheet = workbook.create_sheet(title='All Data')
        
        # Add form title
        all_data_sheet['A1'] = f"{template_type} Form - All Fields"
        all_data_sheet['A1'].font = title_font
        all_data_sheet.merge_cells('A1:C1')
        
        # Add headers
        all_data_sheet['A3'] = "Section"
        all_data_sheet['B3'] = "Field"
        all_data_sheet['C3'] = "Value"
        
        # Style headers
        for cell in all_data_sheet['3:3']:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        # Add all fields from all sections
        row = 4
        for section, fields in complete_form_data.items():
            # Skip empty sections
            if not fields:
                continue
                
            for field, value in fields.items():
                # Add section name
                all_data_sheet[f'A{row}'] = section
                all_data_sheet[f'A{row}'].border = thin_border
                all_data_sheet[f'A{row}'].alignment = left_align
                
                # Add field name
                all_data_sheet[f'B{row}'] = field
                all_data_sheet[f'B{row}'].border = thin_border
                all_data_sheet[f'B{row}'].alignment = left_align
                
                # Add value
                all_data_sheet[f'C{row}'] = value if value is not None else ''
                all_data_sheet[f'C{row}'].border = thin_border
                all_data_sheet[f'C{row}'].alignment = left_align
                
                row += 1
        
        # Set column widths
        all_data_sheet.column_dimensions['A'].width = 25
        all_data_sheet.column_dimensions['B'].width = 30
        all_data_sheet.column_dimensions['C'].width = 50
        
        # Try to save the workbook
        try:
            workbook.save(output)
            logger.debug("Excel file saved successfully")
        except Exception as e:
            logger.error(f"Error saving Excel file: {str(e)}")
            flash("Error generating Excel file", "danger")
            return redirect(url_for('view_form', form_id=form_id))
        
        # Reset file pointer
        output.seek(0)
        
        # Create filename
        file_name = f"{template_type.replace(' ', '_')}_Form_{form_id}.xlsx"
        
        # Return file
        try:
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                download_name=file_name,
                as_attachment=True
            )
        except Exception as e:
            logger.error(f"Error sending Excel file: {str(e)}")
            flash("Error downloading Excel file", "danger")
            return redirect(url_for('view_form', form_id=form_id))
    else:
        flash('Invalid export format.', 'danger')
        return redirect(url_for('view_form', form_id=form_id))

@app.route('/export-all-forms/<format>')
@login_required
def export_all_forms(format):
    """
    Export all forms for the current user, organized by template type
    """
    logger.debug("Starting export-all-forms")
    
    # Get all the user's extracted forms
    extracted_forms = ExtractedForm.query.filter_by(user_id=current_user.id).order_by(ExtractedForm.created_at.desc()).all()
    logger.debug(f"Found {len(extracted_forms)} forms to export")
    
    if not extracted_forms:
        flash('No forms to export.', 'warning')
        return redirect(url_for('dashboard'))
    
    # Group forms by template type
    forms_by_template = {}
    
    for form in extracted_forms:
        template_type = form.template_type
        logger.debug(f"Processing form ID: {form.id}, Template: {template_type}, File: {form.file_name}")
        
        # Initialize entry for this template type if it doesn't exist
        if template_type not in forms_by_template:
            forms_by_template[template_type] = []
            logger.debug(f"Created new template group: {template_type}")
            
        # Get raw form data for debugging
        raw_data = form.extracted_data
        logger.debug(f"Raw data from DB for form {form.id}: {raw_data[:100]}...")
            
        # Get form data
        form_data = form.get_data()
        logger.debug(f"Parsed JSON data for form {form.id}: {str(form_data)[:200]}...")
        
        # Ensure form data follows the template structure
        template = FORM_TEMPLATES.get(template_type, {})
        logger.debug(f"Template sections for {template_type}: {template.keys()}")
        
        complete_form_data = {}
        
        # Process each section
        for section_name, fields_info in template.items():
            section_data = {}
            logger.debug(f"Processing section: {section_name} with {len(fields_info)} fields")
            
            # Initialize with empty data for all fields from template
            for field_info in fields_info:
                field_name = field_info["field"]
                section_data[field_name] = ""
                
            # If section exists in saved data, use those values
            if section_name in form_data and isinstance(form_data[section_name], dict):
                logger.debug(f"Found section {section_name} in form data with {len(form_data[section_name])} fields")
                for field_name, value in form_data[section_name].items():
                    section_data[field_name] = value
                    logger.debug(f"Set field {field_name} = '{value}'")
            else:
                logger.debug(f"Section {section_name} not found in form data or not a dictionary")
                    
            complete_form_data[section_name] = section_data
            
        # Verify data is correctly populated
        logger.debug(f"Processed form {form.id} has {len(complete_form_data)} sections")
        for section, fields in complete_form_data.items():
            logger.debug(f"Section {section} has {len(fields)} fields")
            for field, value in fields.items():
                if value:  # Only log non-empty values to reduce log size
                    logger.debug(f"Field {field} = '{value}'")
            
        # Add form metadata
        form_info = {
            'formId': form.id,
            'fileName': form.file_name,
            'createdAt': form.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'extractedData': complete_form_data
        }
        
        # Add to the template group
        forms_by_template[template_type].append(form_info)
        logger.debug(f"Added form {form.id} to {template_type} group")
    
    # Return the data based on format
    if format == 'json':
        # For client-side Excel export, we'll still provide the JSON
        data = {
            'userId': current_user.id,
            'username': current_user.username,
            'exportDate': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'formsByTemplate': forms_by_template
        }
        
        # Final verification
        for template_type, forms in forms_by_template.items():
            logger.debug(f"Template {template_type} has {len(forms)} forms")
            # Log details of first form as a sample
            if forms:
                sample_form = forms[0]
                logger.debug(f"Sample form ID: {sample_form['formId']}")
                logger.debug(f"Sample form extracted_data sections: {sample_form['extractedData'].keys()}")
                # Log some field values from the first section
                sample_section = list(sample_form['extractedData'].keys())[0]
                logger.debug(f"Sample section {sample_section} data: {sample_form['extractedData'][sample_section]}")
        
        return jsonify(data)
    elif format == 'excel':
        # Server-side Excel generation for all forms
        import io
        from flask import send_file
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Create an in-memory output file
        output = io.BytesIO()
        
        # Create Excel workbook directly with openpyxl for better formatting control
        workbook = openpyxl.Workbook()
        
        # Create the main summary sheet
        summary_sheet = workbook.active
        summary_sheet.title = 'Summary'
        
        # Add title and basic info
        summary_sheet['A1'] = "FormOCR - All Forms Export"
        summary_sheet['A1'].font = Font(size=16, bold=True)
        
        summary_sheet['A2'] = "User:"
        summary_sheet['B2'] = current_user.username
        
        summary_sheet['A3'] = "Export Date:"
        summary_sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Add templates included section
        summary_sheet['A5'] = "Templates Included:"
        summary_sheet['A5'].font = Font(bold=True)
        
        # List template types
        row = 6
        for idx, (template_type, forms) in enumerate(forms_by_template.items()):
            form_count = len(forms)
            summary_sheet[f'A{row}'] = f"{idx + 1}. {template_type} ({form_count} forms)"
            row += 1
        
        # Process each template type
        for template_type, forms in forms_by_template.items():
            if not forms:
                continue  # Skip empty templates
                
            # Create template summary sheet
            safe_template_summary = f"{template_type[:25]}_Summary"[:31].replace('/', '-').replace('\\', '-')
            template_summary_sheet = workbook.create_sheet(title=safe_template_summary)
            
            # Add title
            template_summary_sheet['A1'] = f"{template_type} Forms Summary"
            template_summary_sheet['A1'].font = Font(size=14, bold=True)
            
            # Add headers
            template_summary_sheet['A3'] = "Form ID"
            template_summary_sheet['B3'] = "File Name"
            template_summary_sheet['C3'] = "Created Date"
            
            # Style headers
            for cell in template_summary_sheet['3:3']:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Add forms data
            row = 4
            for form in forms:
                template_summary_sheet[f'A{row}'] = form['formId']
                template_summary_sheet[f'B{row}'] = form['fileName']
                template_summary_sheet[f'C{row}'] = form['createdAt']
                row += 1
            
            # Adjust column widths
            template_summary_sheet.column_dimensions['A'].width = 10
            template_summary_sheet.column_dimensions['B'].width = 30
            template_summary_sheet.column_dimensions['C'].width = 20
            
            # Get template structure from the first form
            template_structure = forms[0]['extractedData']
            
            # Create a detailed combined sheet showing all forms' data 
            safe_template_details = f"{template_type[:20]}_Details"[:31].replace('/', '-').replace('\\', '-')
            details_sheet = workbook.create_sheet(title=safe_template_details)
            
            # Add title
            details_sheet['A1'] = f"{template_type} - All Forms Data"
            details_sheet['A1'].font = Font(size=14, bold=True)
            
            # Add form data
            row = 3
            for form in forms:
                # Add form header
                details_sheet[f'A{row}'] = f"Form ID: {form['formId']}"
                details_sheet[f'B{row}'] = f"File: {form['fileName']}"
                details_sheet[f'C{row}'] = f"Date: {form['createdAt']}"
                details_sheet[f'A{row}'].font = Font(bold=True)
                row += 2
                
                # Add sections and fields
                for section, fields in form['extractedData'].items():
                    # Section header
                    details_sheet[f'A{row}'] = section
                    details_sheet[f'A{row}'].font = Font(bold=True)
                    row += 1
                    
                    # Field headers
                    details_sheet[f'A{row}'] = "Field"
                    details_sheet[f'B{row}'] = "Value"
                    details_sheet[f'A{row}'].font = Font(bold=True)
                    details_sheet[f'B{row}'].font = Font(bold=True)
                    row += 1
                    
                    # Field values
                    for field, value in fields.items():
                        details_sheet[f'A{row}'] = field
                        details_sheet[f'B{row}'] = value or ''
                        row += 1
                    
                    row += 1  # Add space after section
                
                row += 2  # Add space between forms
            
            # Adjust column widths
            details_sheet.column_dimensions['A'].width = 30
            details_sheet.column_dimensions['B'].width = 50
            
            # For each section, create a separate sheet with all forms in one table
            for section_name in template_structure.keys():
                # Get all unique fields across all forms for this section
                all_fields = set()
                for form in forms:
                    form_section = form['extractedData'].get(section_name, {})
                    all_fields.update(form_section.keys())
                
                safe_template = template_type[:12].replace('/', '-').replace('\\', '-')
                safe_section = section_name[:12].replace('/', '-').replace('\\', '-')
                sheet_name = f"{safe_template}-{safe_section}"[:31]
                section_sheet = workbook.create_sheet(title=sheet_name)
                
                # Add title
                section_sheet['A1'] = f"{template_type} - {section_name}"
                section_sheet['A1'].font = Font(size=14, bold=True)
                
                # Add headers for data table
                section_sheet['A3'] = "Form ID"
                section_sheet['B3'] = "File Name" 
                section_sheet['C3'] = "Created Date"
                col = 4
                
                # Import get_column_letter from openpyxl.utils directly
                from openpyxl.utils import get_column_letter
                
                # Add field names as column headers
                for field in sorted(all_fields):
                    column_letter = get_column_letter(col)
                    section_sheet[f'{column_letter}3'] = field
                    col += 1
                
                # Style headers
                for cell in section_sheet['3:3']:
                    if cell.value:  # Only style cells that have values
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                # Add form data
                row = 4
                for form in forms:
                    section_sheet[f'A{row}'] = form['formId']
                    section_sheet[f'B{row}'] = form['fileName']
                    section_sheet[f'C{row}'] = form['createdAt']
                    
                    # Add values for each field
                    col = 4
                    form_section = form['extractedData'].get(section_name, {})
                    for field in sorted(all_fields):
                        column_letter = get_column_letter(col)
                        section_sheet[f'{column_letter}{row}'] = form_section.get(field, '')
                        col += 1
                    
                    row += 1
                
                # Adjust column widths
                section_sheet.column_dimensions['A'].width = 10
                section_sheet.column_dimensions['B'].width = 25
                section_sheet.column_dimensions['C'].width = 20
                
                # Adjust field columns based on field name length
                col = 4
                for field in sorted(all_fields):
                    column_letter = get_column_letter(col)
                    section_sheet.column_dimensions[column_letter].width = min(30, max(15, len(field) * 1.2))
                    col += 1
            
            # Create a consolidated sheet with all sections for this template
            safe_template_consol = template_type[:25].replace('/', '-').replace('\\', '-')
            consol_sheet = workbook.create_sheet(title=safe_template_consol)
            
            # Add title
            consol_sheet['A1'] = f"{template_type} - All Data"
            consol_sheet['A1'].font = Font(size=14, bold=True)
            
            # Add headers
            consol_sheet['A3'] = "Form ID"
            consol_sheet['B3'] = "File Name"
            consol_sheet['C3'] = "Created Date"
            
            # Add section-field headers
            col = 4
            all_headers = []
            
            for section in template_structure.keys():
                section_fields = set()
                for form in forms:
                    form_section = form['extractedData'].get(section, {})
                    section_fields.update(form_section.keys())
                
                # Import get_column_letter if not already imported
                from openpyxl.utils import get_column_letter
                
                for field in sorted(section_fields):
                    column_letter = get_column_letter(col)
                    header = f"{section} - {field}"
                    consol_sheet[f'{column_letter}3'] = header
                    all_headers.append((column_letter, header))
                    col += 1
            
            # Style headers
            for cell in consol_sheet['3:3']:
                if cell.value:  # Only style cells that have values
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Add form data
            row = 4
            for form in forms:
                consol_sheet[f'A{row}'] = form['formId']
                consol_sheet[f'B{row}'] = form['fileName']
                consol_sheet[f'C{row}'] = form['createdAt']
                
                # Add values for each section-field
                for column_letter, header in all_headers:
                    # Parse section and field from header
                    parts = header.split(' - ', 1)
                    if len(parts) == 2:
                        section, field = parts
                        
                        # Get value from form data
                        if section in form['extractedData']:
                            form_section = form['extractedData'][section]
                            value = form_section.get(field, '')
                            consol_sheet[f'{column_letter}{row}'] = value
                
                row += 1
            
            # Adjust column widths
            consol_sheet.column_dimensions['A'].width = 10
            consol_sheet.column_dimensions['B'].width = 25
            consol_sheet.column_dimensions['C'].width = 20
            
            # Adjust field columns based on field name length
            for column_letter, header in all_headers:
                consol_sheet.column_dimensions[column_letter].width = min(20, max(15, len(header) * 0.8))
                
        # Save workbook to output
        workbook.save(output)
        
        # Reset the file pointer
        output.seek(0)
        
        # Create filename
        file_name = f"All_Forms_Export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        # Return the Excel file for download
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name=file_name,
            as_attachment=True
        )
    else:
        flash('Invalid export format.', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/view-saved-forms')
@login_required
def view_saved_forms():
    # Get all the user's extracted forms, ordered by newest first
    extracted_forms = ExtractedForm.query.filter_by(user_id=current_user.id).order_by(ExtractedForm.created_at.desc()).all()
    return render_template('view_saved_forms.html', title='Saved Forms', extracted_forms=extracted_forms)

@app.route('/delete-form/<int:form_id>', methods=['POST'])
@login_required
def delete_form(form_id):
    # Get the form data from the database
    extracted_form = ExtractedForm.query.get_or_404(form_id)
    
    # Check if the form belongs to the current user
    if extracted_form.user_id != current_user.id:
        flash('You do not have permission to delete this form.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Delete the form
    db.session.delete(extracted_form)
    db.session.commit()
    
    flash('Form deleted successfully!', 'success')
    return redirect(url_for('view_saved_forms'))
